from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from prevout import search_prevouts


BTC_DECIMAL = 10**8

def write_save_excel(transactions, bitcoin_address):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = [
      "Date", "Block Height", "Transaction ID", 
      "Sent Amount", "Received Amount", "Received Change"
    ]
    ws.append(headers)
    
    for tx in transactions:
        prevouts = [vin['prevout'] for vin in tx['vin'] if 'prevout' in vin]
        is_in_outputs = any(
          vout['scriptpubkey_address'] == bitcoin_address 
          for vout in tx['vout'] 
          if 'scriptpubkey_address' in vout
        )
        is_in_prevouts = search_prevouts(prevouts, bitcoin_address)
        
        if 'status' in tx and 'block_time' in tx['status'] and 'block_height' in tx['status']:
            tx_time = (
              datetime
              .utcfromtimestamp(tx['status']['block_time'])
              .strftime('%Y-%m-%d %H:%M:%S')
            )
            block_height = tx['status']['block_height']
        else:
            tx_time = "Unconfirmed"
            block_height = "Unconfirmed"

            tx_id = tx['txid']
        
        if is_in_prevouts or is_in_outputs:
            sent_amount = sum(prevout['value'] for prevout in prevouts 
            if prevout.get('scriptpubkey_address') == bitcoin_address) / BTC_DECIMAL
            received_amount = (
              sum(
                  vout['value'] for vout in tx['vout']
                  if vout.get('scriptpubkey_address') == bitcoin_address
              ) / BTC_DECIMAL
            )
            
            net_amount = 0
            received_change = 0
            if is_in_prevouts and is_in_outputs:
                net_amount = sent_amount - received_amount
                received_change = received_amount
            elif is_in_prevouts:
                net_amount = sent_amount
            
            ws.append([tx_time, block_height, tx_id, net_amount, received_amount, received_change])
    
    filename = f"Bitcoin_Transactions_{bitcoin_address}.xlsx"
    wb.save(filename)
    print(f"Transactions saved to {filename}")
    return filename
    pass


def price_log(filepath, transactions):
    print("Logging price data to Excel.")
    wb = load_workbook(filepath)
    if "Price Log" not in wb.sheetnames:
        ws = wb.create_sheet("Price Log")
        ws.append(["Timestamp", "Price USD"])
    else:
        ws = wb["Price Log"]
    
    for transaction in transactions:
        ws.append([transaction['block_time'], transaction['price']])
    
    wb.save(filepath)
    print("Price log updated.")
    pass