import requests
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook

BTC_DECIMAL = 10**8

def is_address_in_prevout(prevouts, bitcoin_address):
    """
    Checks if the bitcoin address is present in the transaction inputs' previous outputs.
    
    Parameters:
    - prevouts (list): The list of previous outputs in transaction inputs.
    - bitcoin_address (str): The Bitcoin address to check.
    
    Returns:
    - bool: True if the address is present in the previous outputs, False otherwise.
    """
    return any(prevout['scriptpubkey_address'] == bitcoin_address for prevout in prevouts if 'scriptpubkey_address' in prevout)

def fetch_transactions(bitcoin_address):
    transactions = []
    base_url = "https://mempool.space/api/address/"
    endpoint = f"{base_url}{bitcoin_address}/txs"
    
    after_txid = None  # Initialize with None for the first request
    while True:
        # If after_txid is set, append it to the endpoint URL to fetch the next page of confirmed transactions
        if after_txid:
            url = f"{endpoint}?after_txid={after_txid}"
        else:
            url = endpoint
        
        response = requests.get(url)
        if response.status_code == 200:
            new_transactions = response.json()
            if not new_transactions:
                break  # Exit the loop if no more transactions are returned
            
            transactions.extend(new_transactions)
            # Update after_txid for the next iteration/page with the last confirmed transaction's txid
            # This assumes the last transaction in the new_transactions list is the oldest, as per the API's sorting
            after_txid = new_transactions[-1]['txid']
        else:
            print(f"Failed to fetch transactions, status code: {response.status_code}")
            break
    
    return transactions


def fetch_price(currency='USD', timestamp=None):
    # print(f"Fetching historical price for {currency} at timestamp: {timestamp}")
    # Using the provided Mempool API endpoint for historical prices.
    api_url = f"https://mempool.space/api/v1/historical-price?currency={currency}&timestamp={timestamp}"
    response = requests.get(api_url)
    if response.status_code == 200:
        data = response.json()
        # Assuming we are interested in the first price entry.
        if data['prices']:
            price_info = data['prices'][0]
            # Return the price in the requested currency.
            return price_info[currency]
        else:
            # print("No price data found.")
            return None
    else:
        # print(f"Failed to fetch price data. Status Code: {response.status_code}")
        return None



def save_transactions_to_excel(transactions, bitcoin_address):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = ["Date", "Block Height", "Transaction ID", "Sent Amount", "Received Amount", "Received Change"]
    ws.append(headers)
    
    for tx in transactions:
        prevouts = [vin['prevout'] for vin in tx['vin'] if 'prevout' in vin]
        is_in_outputs = any(vout['scriptpubkey_address'] == bitcoin_address for vout in tx['vout'] if 'scriptpubkey_address' in vout)
        is_in_prevouts = is_address_in_prevout(prevouts, bitcoin_address)
        
        if 'status' in tx and 'block_time' in tx['status'] and 'block_height' in tx['status']:
            tx_time = datetime.utcfromtimestamp(tx['status']['block_time']).strftime('%Y-%m-%d %H:%M:%S')
            block_height = tx['status']['block_height']
        else:
            tx_time = "Unconfirmed"
            block_height = "Unconfirmed"
        
        tx_id = tx['txid']
        
        if is_in_prevouts or is_in_outputs:
            sent_amount = sum(prevout['value'] for prevout in prevouts if prevout.get('scriptpubkey_address') == bitcoin_address) / BTC_DECIMAL
            received_amount = sum(vout['value'] for vout in tx['vout'] if vout.get('scriptpubkey_address') == bitcoin_address) / BTC_DECIMAL
            
            net_amount = 0
            received_change = 0
            if is_in_prevouts and is_in_outputs:
                net_amount = sent_amount - received_amount
                received_change = received_amount
            elif is_in_prevouts:
                net_amount = sent_amount
            elif is_in_outputs:
                received_amount = received_amount  # This line is redundant, you might want to remove or adjust it.
            
            ws.append([tx_time, block_height, tx_id, net_amount, received_amount, received_change])
    
    filename = f"Bitcoin_Transactions_{bitcoin_address}.xlsx"
    wb.save(filename)
    print(f"Transactions saved to {filename}")
    return filename  # Return the filename to the caller


def price_log(filepath, transactions):
    print("Logging price data to Excel.")
    wb = load_workbook(filepath)
    if "Price Log" not in wb.sheetnames:
        ws = wb.create_sheet("Price Log")
        ws.append(["Timestamp", "Price USD"])
    else:
        ws = wb["Price Log"]
    
    for transaction in transactions:
        # Assuming each transaction is a dictionary with a 'block_time' and 'price' key.
        ws.append([transaction['block_time'], transaction['price']])
    
    wb.save(filepath)
    print("Price log updated.")



def main():
    bitcoin_address = input("\u001b[1m\u001b[4m\u001b[40mPlease enter a Bitcoin address: \u001b[0m")
    transactions = fetch_transactions(bitcoin_address)
    price_data_list = []  # List to collect price data
    
    if transactions:
        for tx in transactions:
            prevouts = [vin['prevout'] for vin in tx['vin'] if 'prevout' in vin]
            is_in_outputs = any(vout['scriptpubkey_address'] == bitcoin_address for vout in tx['vout'] if 'scriptpubkey_address' in vout)
            is_in_prevouts = is_address_in_prevout(prevouts, bitcoin_address)

            # Check if 'status' exists and contains 'block_time' and 'block_height'
            if 'status' in tx and 'block_time' in tx['status'] and 'block_height' in tx['status']:
                tx_time = datetime.utcfromtimestamp(tx['status']['block_time']).strftime('%Y-%m-%d %H:%M:%S')
                block_height = tx['status']['block_height']
            else:
                # Handle unconfirmed transactions by setting a placeholder
                tx_time = "Unconfirmed"
                block_height = "Unconfirmed"
            
            tx_id = tx['txid']

            # Fetch price for the transaction
            if 'status' in tx and 'block_time' in tx['status']:
                price = fetch_price(timestamp=tx['status']['block_time'])
                # Assuming 'price' is directly the USD value. Adjust if it's nested in the response.
                price_data_list.append({'block_time': tx_time, 'price': price})
            
            if is_in_prevouts or is_in_outputs:
                sent_amount = sum(prevout['value'] for prevout in prevouts if prevout.get('scriptpubkey_address') == bitcoin_address) / BTC_DECIMAL
                received_amount = sum(vout['value'] for vout in tx['vout'] if vout.get('scriptpubkey_address') == bitcoin_address) / BTC_DECIMAL

                if is_in_prevouts and is_in_outputs:
                    net_amount = sent_amount - received_amount
                    if net_amount < 0:
                        net_amount_str = f"\u001b[1m{'{:.8f}'.format(net_amount)}\u001b[0m"
                    else:
                        net_amount_str = f"\u001b[33m{'{:.8f}'.format(net_amount)}\u001b[0m"
                    received_amount_str = f"{'{:.8f}'.format(received_amount)}"
                    
                    print(f"\u001b[44;1mDate:\u001b[0m \u001b[40;1m{tx_time}, \u001b[44;1mBlock Height:\u001b[0m \u001b[40;1m{block_height}, \u001b[44;1mTransaction ID:\u001b[0m \u001b[40;1m{tx_id}, \u001b[45;1m\u001b[1mSent Amount: {net_amount_str} \u001b[1m BTC, \u001b[7m\u001b[1mReceived Change: {received_amount_str} BTC\u001b[0m\u001b[1m")

                elif is_in_prevouts:
                    sent_amount_str = f"\u001b[1m\u001b[33m{'{:.8f}'.format(sent_amount)}\u001b[0m" if sent_amount >= 0 else f"\u001b[1m{'{:.8f}'.format(sent_amount)}\u001b[0m"
                    print(f"\u001b[44;1m\u001b[1mDate:\u001b[0m \u001b[40;1m{tx_time}, \u001b[44;1m\u001b[1mBlock Height:\u001b[0m \u001b[40;1m{block_height}, \u001b[44;1m\u001b[1mTransaction ID:\u001b[0m \u001b[40;1m{tx_id}, \u001b[43;1mSent Amount:\u001b[0m {sent_amount_str} \u001b[1m BTC")
                elif is_in_outputs:
                    received_amount_str = f"\u001b[1m\u001b[32m{'{:.8f}'.format(received_amount)}\u001b[0m"
                    print(f"\u001b[44;1m\u001b[1mDate:\u001b[0m \u001b[40;1m{tx_time}, \u001b[44;1m\u001b[1mBlock Height:\u001b[0m \u001b[40;1m{block_height}, \u001b[44;1m\u001b[1mTransaction ID:\u001b[0m \u001b[40;1m{tx_id}, \u001b[42;1mReceived Amount:\u001b[0m {received_amount_str} \u001b[1m BTC")

    # Display the total transaction count with a bright magenta background and bold letters
        print(f"\u001b[45;1mTotal Transaction Count:\u001b[0m \u001b[45;1m{len(transactions)}\u001b[0m")

    # Prompt the user after displaying transactions
        user_choice = input("Do you want to download the displayed transactions as an Excel file? (yes/no): ").lower()
        if user_choice == "yes":
           # save_transactions_to_excel now returns the filename, which we capture
            excel_file_path = save_transactions_to_excel(transactions, bitcoin_address)
            # Pass the filename to price_log
            price_log(excel_file_path, price_data_list)
        else:
            print("Exiting without downloading transactions.")

    else:
        print("No transactions found for this address.")



if __name__ == "__main__":
    main()
